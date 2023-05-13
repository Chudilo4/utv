from openpyxl.workbook import Workbook

from utv_api.models import Worker, TableProject, TableExcel


def create_excel(author_id: int, **kwargs):
    """Создаём excel файл из сформированной таблицы"""
    table = TableProject.objects.get(pk=kwargs['table_pk'])
    wb = Workbook()
    ws = wb.active
    ws['A6'] = 'ЦЕНА для клиента'
    ws['B6'] = table.price_client
    ws['A7'] = 'Плановая себестоимость'
    ws['B7'] = table.planned_cost
    ws['C7'] = 'Фактическая себестоимость'
    ws['D7'] = table.cost
    ws['A8'] = 'Плановая зарплата'
    ws['B8'] = table.planned_salary
    ws['C8'] = 'Фактическая зарплата'
    ws['D8'] = table.salary
    ws['A9'] = 'Плановый гонорар актерам'
    ws['B9'] = table.planned_actors_salary
    ws['C9'] = 'Фактический гонорар актерам'
    ws['D9'] = table.actors_salary
    ws['A10'] = 'Плановые налоги с ФОТ'
    ws['B10'] = table.planned_taxes_FOT
    ws['C10'] = 'Фактические налоги с ФОТ'
    ws['D10'] = table.taxes_FOT
    ws['A11'] = 'Плановая покупка реквизита для организации съемочного процесса/' \
                ' Непредвиденные расходы'
    ws['B11'] = table.planned_other_expenses
    ws['C11'] = 'Фактическая покупка реквизита для организации съемочного процесса/' \
                ' Непредвиденные расходы'
    ws['D11'] = table.other_expenses
    ws['A12'] = 'Плановая покупка музыки'
    ws['B12'] = table.planned_buying_music
    ws['C12'] = 'Фактическа покупка музыки'
    ws['D12'] = table.buying_music
    ws['A13'] = 'Плановые командировочные расходы (суточные, жилье, дорога)'
    ws['B13'] = table.planned_travel_expenses
    ws['C13'] = 'Фактические командировочные расходы (суточные, жилье, дорога)'
    ws['D13'] = table.travel_expenses
    ws['A14'] = 'Плановые такси, транспортные расходы '
    ws['B14'] = table.planned_fare
    ws['C14'] = 'Фактические такси, транспортные расходы '
    ws['D14'] = table.fare
    ws['A15'] = 'Плановые общехозяйственные расходы'
    ws['B15'] = table.planned_general_expenses
    ws['C15'] = 'Фактические общехозяйственные расходы'
    ws['D15'] = table.general_expenses
    ws['A16'] = 'Плановая прибыль'
    ws['B16'] = table.planned_profit
    ws['C16'] = 'Фактическая прибыль'
    ws['D16'] = table.profit
    ws['A17'] = 'Плановая рентабельность'
    ws['B17'] = table.planned_profitability
    ws['C17'] = 'Фактическая рентабельность'
    ws['D17'] = table.profitability
    return wb


def get_my_excels_table(**kwargs):
    """Отдаём пользователю все excel связанные с таблицой"""
    excel = TableExcel.objects.filter(table_id=kwargs['table_pk'])
    return excel


def get_excel(excel_pk: int):
    excel = TableExcel.objects.get(pk=excel_pk)
    return excel


def delete_excel(excel_pk: int):
    excel = get_excel(excel_pk)
    excel.path_excel.delete(save=False)
    excel.delete()


def calculation_table(card_pk, **kwargs):
    """Расчитываем плановые и фактические расчёты затрат и прибыли за проект"""
    # Плановая зарплата сотрудников, Зарплата сотрудников
    planned_salary, salary = salary_executors(
        card_pk=card_pk,
        planned_actors_salary=kwargs.get('planned_actors_salary', 0),
        actors_salary=kwargs.get('actors_salary', 0))
    # Плановые налоги с ФОТ
    planned_taxes_fot = planned_salary * 0.5
    # Налоги с ФОТ
    taxes_fot = salary * 0.5

    # Плановые общехозяйственные расходы
    list_planned_general_expenses = [
        planned_salary, planned_taxes_fot, kwargs.get('planned_other_expenses', 0),
        kwargs.get('planned_buying_music', 0), kwargs.get('planned_travel_expenses', 0),
        kwargs.get('planned_fare', 0)]
    planned_general_expenses = sum(list_planned_general_expenses) * 0.23
    # Общехозяйственные расходы
    list_general_expenses = [
        salary, taxes_fot, kwargs.get('other_expenses', 0), kwargs.get('buying_music', 0),
        kwargs.get('travel_expenses', 0), kwargs.get('fare', 0)]
    general_expenses = sum(list_general_expenses) * 0.23
    # Плановая себестоимость
    list_planned_cost = [
        planned_salary, planned_taxes_fot, kwargs.get('planned_other_expenses', 0),
        planned_general_expenses, kwargs.get('planned_buying_music', 0),
        kwargs.get('planned_travel_expenses', 0), kwargs.get('planned_fare', 0)]
    planned_cost = sum(list_planned_cost)
    # Cебестоимость
    list_cost = [salary, taxes_fot, kwargs.get('other_expenses', 0), general_expenses,
                 kwargs.get('buying_music', 0), kwargs.get('travel_expenses', 0),
                 kwargs.get('fare', 0)]
    cost = sum(list_cost)

    # Плановая прибыль
    planned_profit = kwargs.get('price_client', 15000) - planned_cost
    # Фактическая прибыль
    profit = kwargs.get('price_client', 15000) - cost
    # Плановая рентабельность
    planned_profitability = (planned_profit / kwargs.get('price_client', 15000)) * 100
    # Фактическая рентабельность
    profitability = (profit / kwargs.get('price_client', 15000)) * 100
    return {'planned_salary': planned_salary,
            'salary': salary,
            'planned_taxes_FOT': planned_taxes_fot,
            'taxes_FOT': taxes_fot,
            'planned_general_expenses': planned_general_expenses,
            'general_expenses': general_expenses,
            'planned_cost': planned_cost,
            'cost': cost,
            'planned_profit': planned_profit,
            'profit': profit,
            'planned_profitability': planned_profitability,
            'profitability': profitability
            }


def salary_executors(card_pk, planned_actors_salary: int, actors_salary: int):
    """Расчитываем заработную плату сотрудников за проект и возвращаем её в виде кортежа"""
    # Фактический заработок сотрудников за проект
    workersalary = 0
    # Плановый заработок сотрудников за проект
    planedworkersalary = 0
    for i in Worker.objects.filter(card_id=card_pk):
        for i2 in i.author.employeerate.order_by('-created_time')[:1]:
            planedworkersalary += i2.money * i.scheduled_time
            workersalary += i2.money * i.actual_time
    return planedworkersalary + planned_actors_salary, workersalary + actors_salary
